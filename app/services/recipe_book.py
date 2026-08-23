import io
import re
from decimal import Decimal

import httpx
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from app.models import Item, ItemCategory, MenuItem, RecipeLine, SaleLine, Sauce, SauceLine, Unit, utcnow
from app.services.inventory import apply_pack_stock, make_sku, qty
from app.services.ledger import money
from app.services.sheets import get_or_create_sheet, sheet_export_url

MASTER_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/1Dk1i0OJbTTMrmRaLtYRzb66VKQtkJjYtvYbWaA6T1H8/edit"
)

UNIT_ALIASES = {
    "g": "g",
    "gm": "g",
    "gms": "g",
    "gram": "g",
    "grams": "g",
    "kg": "kg",
    "ml": "ml",
    "ltr": "l",
    "l": "l",
    "pc": "pcs",
    "pcs": "pcs",
    "piece": "pcs",
    "pieces": "pcs",
    "slice": "pcs",
    "pinch": "g",
}

NAME_ALIASES = {
    "onion": "onions",
    "onions": "onions",
    "tomato": "tomatoes",
    "tomatoes": "tomatoes",
    "egg": "eggs",
    "eggs": "eggs",
    "black pepper": "black pepper",
    "blck pepper": "black pepper",
    "cream cheese": "cream cheese",
    "chicken": "chicken thigh",
    "soy sauce": "soy sauce",
    "light soya sauce": "light soya sauce",
    "american musturd": "american musturd",
    "american mustard": "american musturd",
    "avacado": "avacado",
    "avocado": "avacado",
    "cheakpea": "cheakpea",
    "chickpea": "cheakpea",
    "corriander leaves": "corriander leaves",
    "coriander": "corriander leaves",
    "kashmiri lalmirch powder": "kashmiri lalmirch powder",
    "kashmiri chilli": "kashmiri lalmirch powder",
    "kashmiri red chilli": "kashmiri lalmirch powder",
    "red chilli powder": "red chilli powder",
    "spicy red chilli powder": "red chilli powder",
    "chilli flakes": "chilli flakes",
    "chilly flakes": "chilli flakes",
    "french friess": "french friess",
    "fries": "french friess",
    "hokkaido bread": "hokkaido bread",
    "hokaiddo": "hokkaido bread",
    "hokaido bread": "hokkaido bread",
    "hokkaido": "hokkaido bread",
    "s/w bread": "s/w bread",
    "sw bread": "s/w bread",
    "maida/a.p flour": "maida/a.p flour",
    "all purpose flour": "all purpose flour",
    "all pourpose flour": "all purpose flour",
    "all pourpose powder": "all purpose flour",
    "flour": "all purpose flour",
    "process cheese": "process cheese",
    "processed cheese": "process cheese",
    "mozzarella cheese": "mozzarella cheese",
    "mozarella": "mozzarella cheese",
    "mix cheese": "mozzarella cheese",
    "mixed cheese": "mozzarella cheese",
    "mayonnaise": "mayonnaise",
    "mayo": "mayonnaise",
    "kewpie mayo": "mayonnaise",
    "kwipie mayo": "mayonnaise",
    "olive oil": "olive oil",
    "oilve oil": "olive oil",
    "sunflower oil tin": "sunflower oil tin",
    "oil": "sunflower oil tin",
    "vegetable oil": "sunflower oil tin",
    "fresh lemon juice": "fresh lemon juice",
    "lemon juice": "fresh lemon juice",
    "lemon": "lemon",
    "synthetic vinegar": "synthetic vinegar",
    "vinegar": "synthetic vinegar",
    "tomato ketchup": "tomato ketchup",
    "red bellpeppers": "red bellpeppers",
    "red bell pepper": "red bellpeppers",
    "red bellpepper": "red bellpeppers",
    "bell pepper": "red bellpeppers",
    "bell peppers": "red bellpeppers",
    "capsicum": "capsicum",
    "capcisum": "capsicum",
    "green lettuce": "green lettuce",
    "ice burg lettuce": "ice burg lettuce",
    "iceberg lettuce": "ice burg lettuce",
    "lettuce": "green lettuce",
    "jalepenos": "jalepenos",
    "jalapeno": "jalepenos",
    "jalapenos": "jalepenos",
    "kanda lasun masala": "kanda lasun masala",
    "saoji magic paste": "saoji magic paste",
    "cream cheese": "cream cheese",
    "black pepper": "black pepper",
    "honey": "honey",
    "salt": "salt",
    "sugar": "sugar",
    "caster sugar": "caster sugar",
    "castor sugar": "caster sugar",
    "brown sugar": "brown sugar",
    "butter": "butter",
    "garlic": "garlic",
    "ginger": "ginger",
    "basil": "basil",
    "spinach": "spinach",
    "mushroom": "mushroom",
    "curd": "curd",
    "fresh cream": "fresh cream",
    "tahini": "tahini",
    "tahini paste": "tahini",
    "sourdough": "sourdough",
    "chat masala": "chat masala",
    "cumin powder": "cumin powder",
    "cummin powder": "cumin powder",
    "jeera powder": "jeera powder",
    "cumin": "cumin powder",
    "corriander powder": "corriander powder",
    "corn flour": "corn flour",
    "cornflour": "corn flour",
    "corn flower": "corn flour",
    "cornflower": "corn flour",
    "baking powder": "baking powder",
    "baking soda": "baking soda",
    "baking sode": "baking soda",
    "dijon mustard": "dijon mustard",
    "gochujang": "gochujang",
    "smoked peprika": "smoked peprika",
    "smoked red papprika": "smoked peprika",
    "paprika": "smoked peprika",
    "turmeric powder": "turmeric powder",
    "turmeric": "turmeric powder",
    "kadi pata": "kadi pata",
    "kadi patta": "kadi pata",
    "bhut jolokia": "bhut jolokia",
    "bhut jholokia": "bhut jolokia",
    "chicken breast": "chicken breast",
    "chicken thigh": "chicken thigh",
    "chicken ham": "chicken ham",
    "chocolate 46%": "chocolate 46%",
    "dark chocolate(46%)": "chocolate 46%",
    "dark chocolate": "dark chocolate",
    "chocolate 72%": "chocolate 72%",
    "cocoa": "cocoa",
    "hersheys cocoa": "hersheys cocoa",
    "panko bread crumbs": "panko bread crumbs",
    "bread crumbs": "panko bread crumbs",
    "parmesan cheese": "parmesan cheese",
    "cheddar": "process cheese",
    "cheddar cheese": "process cheese",
    "amul cheese": "process cheese",
    "cheese slice": "cheese slice",
    "cherry tomato": "cherry tomato",
    "cucumber": "cucumber",
    "carrot": "carrot",
    "potatoes": "potatoes",
    "potato": "potatoes",
    "green chilli": "green chilli",
    "green chilly": "green chilli",
    "mint leaf": "mint leaf",
    "mint leaves": "mint leaf",
    "parsley": "parsley",
    "peanut": "peanut",
    "walnut": "walnut",
    "wallnuts": "walnut",
    "walniuts": "walnut",
    "sesame (toasted)": "sesame (toasted)",
    "spring onion": "spring onion",
    "whole red chilli": "whole red chilli",
    "bayleaf": "bayleaf",
    "cloves": "cloves",
    "cinnamon powder": "cinnamon powder",
    "cinammon powder": "cinnamon powder",
    "cinamon powder": "cinnamon powder",
    "cinnamon": "cinnamon powder",
    "cinammon": "cinnamon powder",
    "sriraja": "sriraja",
    "sriracha sauce": "sriraja",
    "siracha sauce": "sriraja",
    "hoison sauce": "hoison sauce",
    "kissan jam": "kissan jam",
    "jam": "kissan jam",
    "tomato ketchup": "tomato ketchup",
    "plum sauce": "plum sauce",
    "magic masala": "magic masala",
    "goda masala": "goda masala",
    "gooda masala": "goda masala",
    "garam masala": "garam masala",
    "onion powder": "onion powder",
    "garlic powder": "garlic powder",
    "nutmeg powder": "nutmeg powder",
    "purple cabbage": "purple cabbage",
    "red cabbage": "purple cabbage",
    "yellow bellpeppers": "yellow bellpeppers",
    "green zucckini": "green zucckini",
    "yellow zucckini": "yellow zucckini",
    "beetroot": "beetroot",
    "bhiwapuri chilli": "bhiwapuri chilli",
    "tortilla": "tortilla",
    "croissant": "croissant",
    "chocolate croissant": "chocolate croissant",
    "granola": "granola",
    "kokum agal": "kokum agal",
}

SKIP_TITLES = {
    "yield",
    "s.no.",
    "s.no",
    "ingredient",
    "total cost",
    "per piece",
    "per gram",
    "making price",
    "selling price",
    "portion",
    "portions",
    "pieces",
    "weight",
    "total batch",
    "total weight",
    "nv burger",
    "items",
}

SKIP_INGREDIENTS = {"", "1 serving", "batch", "yellow"}


def _cell(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _norm(value: str) -> str:
    text = re.sub(r"\s+", " ", (value or "").strip().lower())
    text = text.replace("(fine chopped)", "").replace("(minced)", "").replace("(cooked)", "")
    text = re.sub(r"\s+", " ", text).strip(" ,.-")
    return NAME_ALIASES.get(text, text)


def _map_unit(raw: str) -> str:
    key = re.sub(r"[^a-z]", "", (raw or "").lower())
    return UNIT_ALIASES.get(key, "g" if key.startswith("g") else "pcs" if key.startswith("p") else "g")


def _parse_cost(raw: str) -> tuple[Decimal | None, str]:
    text = (raw or "").strip()
    match = re.search(r"([0-9]*\.?[0-9]+)\s*/\s*([a-zA-Z]+)", text)
    if not match:
        return None, "pcs"
    return Decimal(match.group(1)), _map_unit(match.group(2))


def _guess_category(name: str) -> str:
    token = _norm(name)
    if any(word in token for word in ("milk", "cheese", "cream", "curd", "butter", "egg", "yogurt", "delight")):
        return ItemCategory.dairy.value
    if any(word in token for word in ("chicken", "tomato", "onion", "lettuce", "spinach", "mushroom", "avocado", "capsicum", "carrot", "potato", "garlic", "ginger", "chilli", "cabbage", "cucumber", "lemon", "apple", "banana", "basil", "mint")):
        return ItemCategory.produce.value
    if any(word in token for word in ("bread", "bun", "pav", "croissant", "cake", "flour", "maida")):
        return ItemCategory.bakery.value
    if any(word in token for word in ("sauce", "oil", "vinegar", "masala", "powder", "salt", "sugar", "spice")):
        return ItemCategory.dry_goods.value
    if any(word in token for word in ("coffee", "espresso")):
        return ItemCategory.coffee.value
    return ItemCategory.other.value


def _is_header(cells: list[str]) -> bool:
    joined = " ".join(cell.lower() for cell in cells[:6])
    return "ingredient" in joined and "quantity" in joined


def _is_number(value: str) -> bool:
    try:
        Decimal(str(value).replace(",", ""))
        return True
    except Exception:
        return False


def parse_ingri_sheet(workbook) -> list[dict]:
    if "Ingri" not in workbook.sheetnames:
        return []
    items = []
    seen: set[str] = set()
    for row in workbook["Ingri"].iter_rows(values_only=True):
        name = _cell(row[0] if row else "")
        cost = _cell(row[1] if row and len(row) > 1 else "")
        if not name or name.lower() in SKIP_TITLES:
            continue
        key = _norm(name)
        if key in seen:
            continue
        seen.add(key)
        unit_cost, unit = _parse_cost(cost)
        if not cost:
            unit = "pcs"
        items.append({"name": name.strip(), "unit": unit, "unit_cost": unit_cost, "key": key})
    return items


def parse_recipe_sheets(workbook) -> list[dict]:
    recipes: list[dict] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name.lower() == "ingri":
            continue
        current = None
        pending = sheet_name.strip()
        for row in workbook[sheet_name].iter_rows(values_only=True):
            cells = [_cell(value) for value in row]
            if not any(cells):
                continue
            first, second = cells[0], cells[1] if len(cells) > 1 else ""
            labels = {cell.lower() for cell in cells if cell}
            if "selling price" in labels and current is not None:
                for value in reversed(cells):
                    if _is_number(value) and value.lower() not in {"selling price"}:
                        current["price"] = Decimal(value.replace(",", ""))
                        break
                continue
            if labels & {"total cost", "per piece", "per gram", "making price"}:
                continue
            if second.lower() == "yield" or first.lower() == "yield":
                title = pending or (current["name"] + " prep" if current else sheet_name)
                current = {"name": title, "sheet": sheet_name, "lines": [], "price": Decimal("0")}
                recipes.append(current)
                pending = None
                continue
            if _is_header(cells):
                if current is None:
                    current = {
                        "name": pending or sheet_name,
                        "sheet": sheet_name,
                        "lines": [],
                        "price": Decimal("0"),
                    }
                    recipes.append(current)
                    pending = None
                continue
            if (
                first
                and not _is_number(first)
                and first.lower() not in SKIP_TITLES
                and (not second or second.lower() == "yield" or not any(cells[1:4]))
            ):
                pending = first
                continue
            if current is None:
                continue
            if first and (_is_number(first) or re.match(r"^\d+(\.0+)?$", first)) and second:
                name, qty_raw, unit_raw = second, cells[2] if len(cells) > 2 else "", cells[3] if len(cells) > 3 else ""
            else:
                continue
            if _norm(name) in SKIP_INGREDIENTS or name.lower() in SKIP_TITLES:
                continue
            if not _is_number(qty_raw):
                continue
            amount = Decimal(qty_raw.replace(",", ""))
            if amount <= 0:
                continue
            current["lines"].append({"name": name, "qty": amount, "unit": _map_unit(unit_raw or "g")})
    return [recipe for recipe in recipes if recipe["lines"]]


def parse_recipe_workbook(content: bytes) -> tuple[list[dict], list[dict]]:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as error:
        raise HTTPException(status_code=400, detail="Could not read the recipe workbook") from error
    return parse_ingri_sheet(workbook), parse_recipe_sheets(workbook)


def fetch_workbook(url: str) -> bytes:
    export = sheet_export_url(url).replace("format=csv", "format=xlsx").split("&gid=")[0]
    if "export?format=xlsx" not in export:
        match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
        if not match:
            raise HTTPException(status_code=400, detail="That does not look like a Google Sheet link")
        export = f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"
    try:
        response = httpx.get(export, follow_redirects=True, timeout=40.0)
    except httpx.HTTPError as error:
        raise HTTPException(status_code=400, detail="Could not reach Google Sheets") from error
    if response.status_code >= 400 or response.headers.get("content-type", "").startswith("text/html"):
        raise HTTPException(
            status_code=400,
            detail="Could not read the workbook. Share the sheet: Anyone with the link → Viewer.",
        )
    return response.content


def _find_item(db: Session, name: str) -> Item | None:
    key = _norm(name)
    rows = db.scalars(select(Item)).all()
    for item in rows:
        if _norm(item.name) == key:
            return item
    return db.scalar(select(Item).where(func.lower(Item.name) == name.strip().lower()))


PREP_EXACT = {
    "kheema",
    "cream",
    "cream base",
    "hummus",
    "falafel",
    "salsa",
    "sourcream",
    "sour cream",
    "pesto",
    "poolish",
    "garlic aoli",
    "garlic aioli",
    "aioli sauce",
    "pasta sauce",
    "spuli sauce",
    "cc sauce",
    "chilly oil",
    "chilli oil",
    "chili oil",
    "cake sponge",
    "pancake batter",
    "esp cream",
    "salted caramel",
    "strussel",
    "caramalised apple",
    "chicken marination",
}

PREP_SUFFIXES = (
    " sauce",
    " aioli",
    " aoli",
    " dressing",
    " marinade",
    " marination",
    " batter",
    " sponge",
    " chutney",
    " salsa",
    " pesto",
    " poolish",
    " dip",
    " oil",
    " prep",
)


def is_prep_recipe(name: str) -> bool:
    label = re.sub(r"\s+", " ", (name or "").strip().lower())
    if not label:
        return False
    if label in PREP_EXACT:
        return True
    return any(label.endswith(suffix) for suffix in PREP_SUFFIXES)


def _set_sauce_lines(db: Session, sauce: Sauce, pairs: list[tuple[int, Decimal]]) -> None:
    sauce.lines.clear()
    db.flush()
    merged: dict[int, Decimal] = {}
    for item_id, quantity in pairs:
        merged[item_id] = merged.get(item_id, Decimal("0")) + qty(quantity)
    for item_id, quantity in merged.items():
        db.add(SauceLine(sauce_id=sauce.id, item_id=item_id, quantity=quantity))


def upsert_sauce_from_lines(db: Session, name: str, pairs: list[tuple[int, Decimal]]) -> tuple[Sauce, bool]:
    sauce = db.scalar(select(Sauce).where(func.lower(Sauce.name) == name.strip().lower()))
    created = sauce is None
    if created:
        sauce = Sauce(name=name.strip(), active=True)
        db.add(sauce)
        db.flush()
    _set_sauce_lines(db, sauce, pairs)
    return sauce, created


def split_prep_menu_items(db: Session) -> int:
    moved = 0
    rows = db.scalars(select(MenuItem).options(selectinload(MenuItem.recipe_lines))).all()
    for menu_item in rows:
        if not is_prep_recipe(menu_item.name):
            continue
        sold = db.scalar(select(SaleLine.id).where(SaleLine.menu_item_id == menu_item.id).limit(1))
        if sold:
            continue
        pairs = [(line.item_id, qty(line.quantity)) for line in menu_item.recipe_lines if line.item_id]
        if pairs:
            upsert_sauce_from_lines(db, menu_item.name, pairs)
        db.delete(menu_item)
        moved += 1
    return moved


def ensure_ingredient(db: Session, name: str, unit: str, unit_cost: Decimal | None = None) -> tuple[Item, bool]:
    item = _find_item(db, name)
    if item:
        if unit_cost and qty(item.unit_cost) == 0:
            item.unit_cost = money(unit_cost)
        return item, False
    stock_unit = unit if unit in {item.value for item in Unit} else "g"
    item = Item(
        sku=make_sku(db, name),
        name=name.strip(),
        category=_guess_category(name),
        unit=stock_unit,
        serving_size=Decimal("1"),
        serving_unit=stock_unit,
        active=True,
    )
    apply_pack_stock(item, qty_per_unit=1, units_on_hand=0, total_price=0)
    if unit_cost:
        item.unit_cost = money(unit_cost)
    db.add(item)
    db.flush()
    return item, True


def import_recipe_book(db: Session, url: str | None = None, content: bytes | None = None) -> dict:
    sheet = get_or_create_sheet(db, "recipe_book")
    source = (url or sheet.url or MASTER_SHEET_URL).strip()
    payload = content if content is not None else fetch_workbook(source)
    ingredients, recipes = parse_recipe_workbook(payload)
    created_items = updated_items = created_recipes = updated_recipes = created_sauces = updated_sauces = 0
    for raw in ingredients:
        _item, created = ensure_ingredient(db, raw["name"], raw["unit"], raw["unit_cost"])
        created_items += int(created)
        updated_items += int(not created)
    for recipe in recipes:
        for line in recipe["lines"]:
            _item, created = ensure_ingredient(db, line["name"], line["unit"])
            created_items += int(created)
    for recipe in recipes:
        pairs = []
        for line in recipe["lines"]:
            ingredient = _find_item(db, line["name"])
            if ingredient is None:
                continue
            pairs.append((ingredient.id, qty(line["qty"])))
        if is_prep_recipe(recipe["name"]):
            _sauce, created = upsert_sauce_from_lines(db, recipe["name"], pairs)
            created_sauces += int(created)
            updated_sauces += int(not created)
            continue
        menu_item = db.scalar(
            select(MenuItem)
            .options(selectinload(MenuItem.recipe_lines))
            .where(func.lower(MenuItem.name) == recipe["name"].strip().lower())
        )
        if menu_item is None:
            menu_item = MenuItem(
                name=recipe["name"].strip(),
                category="food",
                price=money(recipe.get("price") or 0),
                active=True,
            )
            db.add(menu_item)
            db.flush()
            created_recipes += 1
        else:
            menu_item.price = money(recipe.get("price") or menu_item.price)
            menu_item.recipe_lines.clear()
            db.flush()
            updated_recipes += 1
        for line in recipe["lines"]:
            ingredient = _find_item(db, line["name"])
            if ingredient is None:
                continue
            unit = line["unit"] if line["unit"] in {item.value for item in Unit} else ingredient.unit
            db.add(
                RecipeLine(
                    menu_item_id=menu_item.id,
                    item_id=ingredient.id,
                    quantity=qty(line["qty"]),
                    unit=unit,
                )
            )
    moved = split_prep_menu_items(db)
    sheet.url = source
    sheet.last_synced_at = utcnow()
    sheet.last_message = (
        f"Ingredients +{created_items}. Dishes +{created_recipes} / {updated_recipes} updated. "
        f"Sauces +{created_sauces} / {updated_sauces} updated"
        + (f", moved {moved} from dishes" if moved else "")
        + ". Stock left as-is."
    )
    db.commit()
    return {
        "created": created_items + created_recipes,
        "updated": updated_items + updated_recipes,
        "skipped": 0,
        "errors": [],
        "last_synced_at": sheet.last_synced_at,
        "last_message": sheet.last_message,
        "url": sheet.url,
        "ingredients_created": created_items,
        "recipes_created": created_recipes,
        "recipes_updated": updated_recipes,
    }
