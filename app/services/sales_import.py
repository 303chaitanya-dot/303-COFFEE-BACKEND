import csv
import hashlib
import io
import json
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models import MenuItem, PetPoojaOrder
from app.services.inventory import record_sale
from app.services.petpooja import resolve_menu_item

SKIP = {
    "total",
    "min",
    "min.",
    "max",
    "max.",
    "avg",
    "avg.",
    "sub total",
    "subtotal",
    "category",
    "grand total",
}


def _norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _cell(row: list, index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]


def read_rows(filename: str, content: bytes) -> list[list]:
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        text = content.decode("utf-8-sig", errors="replace")
        return [list(row) for row in csv.reader(io.StringIO(text))]
    if name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise HTTPException(status_code=400, detail="Excel support is missing") from error
        book = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet = book.active
        return [[cell if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)]
    if name.endswith(".xls"):
        raise HTTPException(status_code=400, detail="Save the Pet Pooja export as .xlsx or .csv and upload that")
    if name.endswith(".numbers"):
        raise HTTPException(status_code=400, detail="Export from Numbers as CSV or Excel (.xlsx), then upload that file")
    raise HTTPException(status_code=400, detail="Upload the Item Wise Sales Report as .xlsx or .csv")


def parse_report_date(rows: list[list]) -> date | None:
    for row in rows[:8]:
        labels = [_norm(cell) for cell in row]
        if labels and labels[0].startswith("date"):
            raw = " ".join(str(cell) for cell in row[1:] if cell)
            match = re.search(r"(\d{4}-\d{2}-\d{2})", raw)
            if match:
                return date.fromisoformat(match.group(1))
            match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", raw)
            if match:
                return date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
    return None


def parse_item_wise_rows(rows: list[list]) -> list[dict]:
    header_index = None
    mapping: dict[str, int] = {}
    for index, row in enumerate(rows):
        labels = [_norm(cell) for cell in row]
        if any(label in {"item", "item name", "itemname"} for label in labels) and any(
            label.startswith("qty") or label == "quantity" for label in labels
        ):
            header_index = index
            mapping = {label: i for i, label in enumerate(labels) if label}
            break
    if header_index is None:
        raise HTTPException(
            status_code=400,
            detail="Could not find Item and Qty columns. Upload Pet Pooja's Item Wise Sales Report.",
        )

    def idx(*names) -> int | None:
        for name in names:
            if name in mapping:
                return mapping[name]
        return None

    item_i = idx("item", "item name", "itemname")
    code_i = idx("code", "item code")
    qty_i = idx("qty", "qty.", "quantity")
    items = []
    for row in rows[header_index + 1 :]:
        first = _norm(_cell(row, 0))
        name = str(_cell(row, item_i) or "").strip()
        if first in SKIP and not name:
            continue
        if not name:
            continue
        raw_qty = _cell(row, qty_i)
        try:
            quantity = Decimal(str(raw_qty).replace(",", ""))
        except (InvalidOperation, TypeError):
            continue
        if quantity <= 0:
            continue
        items.append(
            {
                "name": name,
                "code": str(_cell(row, code_i) or "").strip() or None,
                "quantity": quantity,
            }
        )
    if not items:
        raise HTTPException(status_code=400, detail="No sold items found in that report")
    return items


def apply_item_wise_report(db: Session, filename: str, content: bytes) -> dict:
    rows = read_rows(filename, content)
    report_date = parse_report_date(rows) or date.today()
    items = parse_item_wise_rows(rows)
    digest = hashlib.sha256(content).hexdigest()
    external_id = f"itemwise:{report_date.isoformat()}:{digest[:16]}"
    existing = db.scalar(select(PetPoojaOrder).where(PetPoojaOrder.external_order_id == external_id))
    if existing and existing.sale_id:
        raise HTTPException(status_code=409, detail="This report was already applied")

    sale_lines = []
    applied = []
    skipped = []
    for item in items:
        qty = int(item["quantity"])
        menu_item = resolve_menu_item(db, item["code"], item["name"])
        if menu_item is None:
            skipped.append({"name": item["name"], "quantity": qty, "reason": "No matching dish"})
            continue
        menu_item = db.scalar(
            select(MenuItem).options(selectinload(MenuItem.recipe_lines)).where(MenuItem.id == menu_item.id)
        )
        if not menu_item or not menu_item.recipe_lines:
            skipped.append({"name": item["name"], "quantity": qty, "reason": "Dish has no recipe yet"})
            continue
        sale_lines.append({"menu_item_id": menu_item.id, "quantity": qty})
        applied.append({"name": item["name"], "quantity": qty, "dish": menu_item.name})

    record = existing or PetPoojaOrder(external_order_id=external_id)
    record.raw_payload = json.dumps({"filename": filename, "date": report_date.isoformat(), "digest": digest})
    record.unmapped_json = json.dumps(skipped) if skipped else None
    record.status = "unmapped" if not sale_lines else "applied"
    if existing is None:
        db.add(record)
    db.flush()

    if sale_lines:
        sale = record_sale(
            db,
            payment_method="other",
            notes=f"Pet Pooja item-wise {report_date.isoformat()} ({Path(filename).name})",
            sold_at=datetime.combine(report_date, datetime.min.time()),
            lines=sale_lines,
        )
        record.sale_id = sale.id
        record.status = "partial" if skipped else "applied"
    db.flush()
    return {
        "filename": filename,
        "report_date": report_date.isoformat(),
        "applied": applied,
        "skipped": skipped,
        "sale_id": record.sale_id,
        "status": record.status,
        "message": f"Applied {len(applied)} dishes, skipped {len(skipped)}",
    }
